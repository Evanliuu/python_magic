# -*- coding:utf-8 -*-
import xlrd
import xlwt
import openpyxl

__author__ = 'Evan'


def write_excel_table(write_info, table_name='excel_example.xls', sheet_name='first_page'):
    """
    写入Excel表格
    :param write_info: 要写入Excel表格的数据
    :param table_name: Excel表格名称
    :param sheet_name: Excel页面名称
    :return:
    """
    # 创建一个Excel文档对象
    ex_wt = xlwt.Workbook()
    # 添加一个新的工作表
    sheet1 = ex_wt.add_sheet(sheet_name, cell_overwrite_ok=True)

    for row_index, each_row in enumerate(write_info):
        if isinstance(each_row, (list, tuple)):
            # 如果是列表或者元组，循环写入每条数据
            for column_index, each_column in enumerate(each_row):
                sheet1.write(row_index, column_index, each_column)
        else:
            # 写入一条数据
            sheet1.write(row_index, 0, each_row)
    # 保存数据到表格
    ex_wt.save(table_name)


def read_excel_table_color(file_name):
    """
    读取Excel表格，返回每个单元格的颜色和值
    :param file_name: Excel文档名称
    :return:
    """
    workbook = openpyxl.load_workbook(file_name)  # 打开Excel文档
    worksheet = workbook.active  # 读取Excel表格
    rows, cols = worksheet.max_row, worksheet.max_column  # 返回表格内的总行数和总列数
    result = []
    for row in range(1, rows + 1):  # 循环遍历所有行
        row_data_total = []
        for column in range(1, cols + 1):  # 循环遍历所有列
            ce = worksheet.cell(row=row, column=column)
            row_data_total.append([ce.fill.start_color.rgb, ce.value])  # 返回每个单元格的颜色和值
        result.append(row_data_total)
    return result


def read_excel_table(file_name, sheet_index=0):
    """
    读取Excel表格
    :param file_name: Excel表格名称
    :param sheet_index: 表格的页面索引值，第一页为0，以此类推
    :return:
    """
    # 打开Excel文档
    ex_rd = xlrd.open_workbook(filename=file_name)
    # 读取Excel的表格（sheet_index=0 读取第一张表格）
    sheet = ex_rd.sheet_by_index(sheet_index)

    result = []
    for i in range(sheet.nrows):
        # 循环读取表格每行数据
        row_data = sheet.row_values(i)
        result.append(row_data)
    return result


if __name__ == '__main__':
    message = [
        ['name', 'id'],
        ['evan', '66'],
        'writer finish'
    ]
    # 写入Excel表
    write_excel_table(write_info=message)
    # 读取Excel表
    print(read_excel_table(file_name='excel_example.xls'))
